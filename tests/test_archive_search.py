from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook as real_load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
APP_DIR = PROJECT_ROOT / "app"
for path in (APP_DIR, PROJECT_ROOT):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from beeline_issue_tracker.data.archive import ExcelArchive
from beeline_issue_tracker.data.archive_search import search_excel_archive, search_result_dedupe_key
from beeline_issue_tracker.domain import IssueSearchResult, ResolvedIssue


class ArchiveSearchTest(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.archive_path = self.root / "archive.xlsx"
        self.resolved = ResolvedIssue(
            id=1,
            original_issue_id=101,
            public_issue_id="ISS-20260603-001",
            machine_number="DEMO-101",
            logged_by="Operator",
            title="Robot vacuum cup fault",
            description="Robot could not confirm vacuum.",
            severity="Non-Critical",
            category="Automation",
            created_at="2026-06-03T08:00:00+00:00",
            resolved_at="2026-06-03T08:30:00+00:00",
            resolved_by="Technician",
            solution="Replaced worn vacuum cup.",
            archive_status="archived",
            archive_error="",
        )
        ExcelArchive(self.archive_path).append_resolved_issue(self.resolved)

    def tearDown(self) -> None:
        self.tmp.cleanup()

    def test_deep_search_reads_excel_read_only_and_data_only(self) -> None:
        with patch("beeline_issue_tracker.data.archive_search.load_workbook", wraps=real_load_workbook) as load:
            results = search_excel_archive(self.archive_path, query="vacuum cup", limit=10)

        self.assertEqual(["Excel Archive"], [result.source for result in results])
        self.assertEqual(["Robot vacuum cup fault"], [result.title for result in results])
        self.assertTrue(load.call_args.kwargs["read_only"])
        self.assertTrue(load.call_args.kwargs["data_only"])

    def test_deep_search_skips_records_already_returned_by_sqlite(self) -> None:
        quick_result = IssueSearchResult(
            state="resolved",
            source="Recent Archive",
            issue_id=1,
            public_issue_id="ISS-20260603-001",
            machine_number="DEMO-101",
            machine_name="Demo Machine",
            machine_model="",
            title="Robot vacuum cup fault",
            description="Robot could not confirm vacuum.",
            status="Non-Critical",
            category="Automation",
            logged_by="Operator",
            created_at="2026-06-03T08:00:00+00:00",
            updated_at="2026-06-03T08:30:00+00:00",
            resolved_at="2026-06-03T08:30:00+00:00",
            resolved_by="Technician",
            resolution="Replaced worn vacuum cup.",
            history_text="",
        )

        results = search_excel_archive(
            self.archive_path,
            query="vacuum cup",
            existing_keys={search_result_dedupe_key(quick_result)},
            limit=10,
        )

        self.assertEqual([], results)


if __name__ == "__main__":
    unittest.main()
