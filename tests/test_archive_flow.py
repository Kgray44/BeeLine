from __future__ import annotations

import tempfile
import unittest
from datetime import datetime, timezone
from pathlib import Path
import sys

PROJECT_ROOT = Path(__file__).resolve().parents[1]
APP_DIR = PROJECT_ROOT / "app"
for path in (APP_DIR, PROJECT_ROOT):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from openpyxl import load_workbook

from beeline_issue_tracker.data import archive as archive_module
from beeline_issue_tracker.data.archive import (
    ARCHIVE_SHEET,
    GROUPED_HEADERS,
    GROUPED_SHEET,
    HEADERS,
    INFO_SHEET,
    ExcelArchive,
    refresh_archive_workbook,
)
from beeline_issue_tracker.data.archive_worker import ArchiveIssueTask, ArchiveRetryTask
from beeline_issue_tracker.data.database import initialize_database
from beeline_issue_tracker.data.repository import IssueRepository
from beeline_issue_tracker.domain import LINE_DOWN, NON_CRITICAL, ResolvedIssue, display_issue_id


DEMO_MACHINES = (
    ("DEMO-101", "Demo Molder 101", "Demo Hive", "Cell A", "DEMO-ASSET-101", 10),
)


class ArchiveFlowTest(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.db_path = self.root / "beeline.sqlite3"
        self.archive_path = self.root / ".archive" / "beeline_resolved_archive.xlsx"
        initialize_database(self.db_path, DEMO_MACHINES)
        self.repository = IssueRepository(self.db_path)
        self.machine_number = self.repository.list_machines_with_status()[0].machine_number

    def tearDown(self) -> None:
        self.tmp.cleanup()

    def test_resolve_moves_issue_to_cache_and_appends_excel_archive(self) -> None:
        issue = self.repository.log_issue(
            machine_number=self.machine_number,
            logged_by="Archive Test",
            title="Nozzle temp drift",
            description="Temperature is drifting outside the expected band.",
            severity=NON_CRITICAL,
            category="Process",
        )

        resolved = self.repository.resolve_issue(
            issue.id,
            solution="Adjusted PID settings and verified stable temperature.",
            resolved_by="Archive Test",
        )
        result = ExcelArchive(self.archive_path).append_resolved_issue(resolved)
        self.repository.mark_archive_result(resolved.id, success=True)

        self.assertEqual([], self.repository.list_active_issues(self.machine_number))
        recent = self.repository.list_recent_resolved_issues(self.machine_number)
        self.assertEqual(1, len(recent))
        self.assertEqual("archived", recent[0].archive_status)
        self.assertTrue(self.archive_path.exists())
        self.assertEqual(2, result.row_number)
        self.assertFalse(result.grouped_refresh_deferred)

        workbook = load_workbook(self.archive_path, data_only=True)
        self.assertIn(ARCHIVE_SHEET, workbook.sheetnames)
        self.assertIn(GROUPED_SHEET, workbook.sheetnames)
        worksheet = workbook[ARCHIVE_SHEET]
        self.assertEqual("visible", worksheet.sheet_state)
        self.assertEqual(HEADERS, tuple(cell.value for cell in worksheet[1][: len(HEADERS)]))
        self.assertEqual(resolved.id, worksheet["A2"].value)
        self.assertEqual(issue.id, worksheet["B2"].value)
        self.assertEqual(display_issue_id(resolved), worksheet["C2"].value)
        self.assertEqual(self.machine_number, worksheet["D2"].value)
        self.assertEqual("Nozzle temp drift", worksheet["F2"].value)
        self.assertEqual("Adjusted PID settings and verified stable temperature.", worksheet["M2"].value)

        grouped = workbook[GROUPED_SHEET]
        self.assertEqual(GROUPED_HEADERS, tuple(cell.value for cell in grouped[3][: len(GROUPED_HEADERS)]))
        self.assertTrue(any(grouped.cell(row=row, column=1).value for row in range(4, grouped.max_row + 1)))
        self.assertEqual(display_issue_id(resolved), grouped.cell(row=5, column=2).value)
        self.assertEqual(1, grouped.row_dimensions[5].outlineLevel)

    def test_archive_append_adds_rows_without_overwriting_existing_rows(self) -> None:
        first = self._resolve_and_archive("First issue")
        second = self._resolve_and_archive("Second issue")

        workbook = load_workbook(self.archive_path, data_only=True)
        worksheet = workbook[ARCHIVE_SHEET]
        self.assertEqual(3, worksheet.max_row)
        self.assertEqual(first.id, worksheet["A2"].value)
        self.assertEqual(second.id, worksheet["A3"].value)

    def test_grouped_sheet_sorts_dates_and_rows_newest_first_without_duplicate_ids(self) -> None:
        older_morning = self._resolved_issue(
            1,
            title="Older morning issue",
            severity=NON_CRITICAL,
            created_at="2026-06-01T07:00:00+00:00",
            resolved_at="2026-06-01T08:00:00+00:00",
        )
        older_noon = self._resolved_issue(
            2,
            title="Older noon issue",
            severity=LINE_DOWN,
            created_at="2026-06-01T09:00:00+00:00",
            resolved_at="2026-06-01T12:00:00+00:00",
        )
        newest = self._resolved_issue(
            3,
            title="Newest day issue",
            severity=NON_CRITICAL,
            created_at="2026-06-02T06:15:00+00:00",
            resolved_at="2026-06-02T07:30:00+00:00",
        )

        archive = ExcelArchive(self.archive_path)
        archive.append_resolved_issue(older_morning)
        archive.append_resolved_issue(older_noon)
        archive.append_resolved_issue(newest)
        archive.append_resolved_issue(older_noon)

        workbook = load_workbook(self.archive_path, data_only=True)
        raw = workbook[ARCHIVE_SHEET]
        self.assertEqual(4, raw.max_row)
        self.assertEqual([1, 2, 3], [raw.cell(row=row, column=1).value for row in range(2, 5)])

        grouped = workbook[GROUPED_SHEET]
        date_headers = [
            (row, grouped.cell(row=row, column=1).value)
            for row in range(4, grouped.max_row + 1)
            if grouped.cell(row=row, column=1).value in {"2026-06-02", "2026-06-01"}
        ]
        self.assertEqual(["2026-06-02", "2026-06-01"], [value for _row, value in date_headers])

        june_2_row = date_headers[0][0]
        june_1_row = date_headers[1][0]
        self.assertEqual("103", grouped.cell(row=june_2_row + 1, column=2).value)
        self.assertEqual(["102", "101"], [grouped.cell(row=june_1_row + offset, column=2).value for offset in (1, 2)])
        self.assertEqual(1, grouped.row_dimensions[june_2_row + 1].outlineLevel)
        self.assertEqual(1, grouped.row_dimensions[june_1_row + 1].outlineLevel)
        self.assertEqual(1, grouped.row_dimensions[june_1_row + 2].outlineLevel)

    def test_large_archive_defers_grouped_refresh_until_repair(self) -> None:
        old_threshold = archive_module.GROUPED_REFRESH_ROW_THRESHOLD
        archive_module.GROUPED_REFRESH_ROW_THRESHOLD = 2
        try:
            archive = ExcelArchive(self.archive_path)
            first = archive.append_resolved_issue(self._resolved_issue(
                1,
                title="First issue",
                severity=NON_CRITICAL,
                created_at="2026-06-01T07:00:00+00:00",
                resolved_at="2026-06-01T08:00:00+00:00",
            ))
            second = archive.append_resolved_issue(self._resolved_issue(
                2,
                title="Second issue",
                severity=NON_CRITICAL,
                created_at="2026-06-01T09:00:00+00:00",
                resolved_at="2026-06-01T10:00:00+00:00",
            ))
            third = archive.append_resolved_issue(self._resolved_issue(
                3,
                title="Third issue",
                severity=LINE_DOWN,
                created_at="2026-06-02T09:00:00+00:00",
                resolved_at="2026-06-02T10:00:00+00:00",
            ))

            self.assertFalse(first.grouped_refresh_deferred)
            self.assertFalse(second.grouped_refresh_deferred)
            self.assertTrue(third.grouped_refresh_deferred)

            workbook = load_workbook(self.archive_path, data_only=True)
            info_values = [cell.value for row in workbook[INFO_SHEET].iter_rows(values_only=False) for cell in row]
            self.assertIn("Deferred; run python run_beeline.py --repair-archive", info_values)

            refresh_archive_workbook(self.archive_path)
            repaired = load_workbook(self.archive_path, data_only=True)
            repaired_info = [cell.value for row in repaired[INFO_SHEET].iter_rows(values_only=False) for cell in row]
            self.assertIn("Current", repaired_info)
            self.assertIn(GROUPED_SHEET, repaired.sheetnames)
        finally:
            archive_module.GROUPED_REFRESH_ROW_THRESHOLD = old_threshold

    def test_failed_excel_archive_marks_resolved_issue_failed_in_sqlite(self) -> None:
        issue = self.repository.log_issue(
            machine_number=self.machine_number,
            logged_by="Archive Test",
            title="Guard fault",
            description="Guard switch faulted during cycle start.",
            severity=LINE_DOWN,
            category="Safety",
        )
        resolved = self.repository.resolve_issue(
            issue.id,
            solution="Replaced guard switch.",
            resolved_by="Archive Test",
        )

        blocked_archive_path = self.root / "blocked_archive.xlsx"
        blocked_archive_path.mkdir()
        task = ArchiveIssueTask(blocked_archive_path, self.repository, resolved)
        with self.assertLogs("beeline_issue_tracker.data.archive_worker", level="ERROR"):
            task.run()

        self.assertEqual([], self.repository.list_active_issues(self.machine_number))
        recent = self.repository.list_recent_resolved_issues(self.machine_number)
        self.assertEqual(1, len(recent))
        self.assertEqual("failed", recent[0].archive_status)
        self.assertTrue(recent[0].archive_error)

    def test_cache_trimming_preserves_pending_and_failed_archive_writes(self) -> None:
        old_archived = self._resolve_and_archive("Old archived")
        newest_archived = self._resolve_and_archive("Newest archived")
        pending = self._resolve_without_archive("Old pending")
        failed = self._resolve_without_archive("Old failed")
        self.repository.mark_archive_result(failed.id, success=False, error="workbook unavailable")
        self._set_resolved_at(old_archived.id, "2025-01-01T08:00:00+00:00")
        self._set_resolved_at(newest_archived.id, "2026-05-01T08:00:00+00:00")
        self._set_resolved_at(pending.id, "2025-01-02T08:00:00+00:00")
        self._set_resolved_at(failed.id, "2025-01-03T08:00:00+00:00")

        deleted = self.repository.trim_resolved_issue_cache(
            keep_days=30,
            keep_minimum=1,
            keep_per_machine_minimum=0,
            now=datetime(2026, 6, 3, tzinfo=timezone.utc),
        )

        self.assertEqual(1, deleted)
        self.assertIsNone(self.repository.get_resolved_issue(old_archived.id))
        self.assertIsNotNone(self.repository.get_resolved_issue(newest_archived.id))
        self.assertEqual("pending", self.repository.get_resolved_issue(pending.id).archive_status)
        self.assertEqual("failed", self.repository.get_resolved_issue(failed.id).archive_status)

    def test_retry_failed_archive_writes_appends_excel_and_marks_archived(self) -> None:
        failed = self._resolve_without_archive("Retry archived")
        self.repository.mark_archive_result(failed.id, success=False, error="locked")

        ArchiveRetryTask(self.archive_path, self.repository).run()

        retried = self.repository.get_resolved_issue(failed.id)
        self.assertIsNotNone(retried)
        assert retried is not None
        self.assertEqual("archived", retried.archive_status)
        workbook = load_workbook(self.archive_path, data_only=True)
        ids = [row[0] for row in workbook[ARCHIVE_SHEET].iter_rows(min_row=2, values_only=True)]
        self.assertIn(failed.id, ids)

    def _resolve_and_archive(self, title: str):
        issue = self.repository.log_issue(
            machine_number=self.machine_number,
            logged_by="Archive Test",
            title=title,
            description=f"{title} description.",
            severity=NON_CRITICAL,
            category="Test",
        )
        resolved = self.repository.resolve_issue(
            issue.id,
            solution=f"{title} fix.",
            resolved_by="Archive Test",
        )
        ExcelArchive(self.archive_path).append_resolved_issue(resolved)
        self.repository.mark_archive_result(resolved.id, success=True)
        return resolved

    def _resolve_without_archive(self, title: str):
        issue = self.repository.log_issue(
            machine_number=self.machine_number,
            logged_by="Archive Test",
            title=title,
            description=f"{title} description.",
            severity=NON_CRITICAL,
            category="Test",
        )
        return self.repository.resolve_issue(
            issue.id,
            solution=f"{title} fix.",
            resolved_by="Archive Test",
        )

    def _set_resolved_at(self, resolved_id: int, resolved_at: str) -> None:
        with self.repository_connection() as conn:
            conn.execute(
                """
                UPDATE resolved_issues_cache
                SET resolved_at = ?
                WHERE id = ?
                """,
                (resolved_at, resolved_id),
            )

    def repository_connection(self):
        from beeline_issue_tracker.data.database import connect

        return connect(self.repository.db_path)

    def _resolved_issue(
        self,
        issue_id: int,
        *,
        title: str,
        severity: str,
        created_at: str,
        resolved_at: str,
    ) -> ResolvedIssue:
        return ResolvedIssue(
            id=issue_id,
            original_issue_id=issue_id + 100,
            machine_number=self.machine_number,
            logged_by="Archive Test",
            title=title,
            description=f"{title} description.",
            severity=severity,
            category="Test",
            created_at=created_at,
            resolved_at=resolved_at,
            resolved_by="Archive Test",
            solution=f"{title} fix.",
            archive_status="pending",
            archive_error="",
        )


if __name__ == "__main__":
    unittest.main()
