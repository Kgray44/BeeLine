from __future__ import annotations

import logging
from collections.abc import Callable, Iterable
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from PySide6.QtCore import QObject, QRunnable, Signal, Slot

from beeline_issue_tracker.data.archive import ARCHIVE_SHEET
from beeline_issue_tracker.domain import IssueSearchResult
from beeline_issue_tracker.perf import elapsed_ms, log as perf_log, now as perf_now


logger = logging.getLogger(__name__)

QUICK_SEARCH_LIMIT = 50
DEEP_SEARCH_LIMIT = 250
DEEP_SEARCH_HARD_LIMIT = 1000
EXCEL_BATCH_SIZE = 25

SEARCH_FIELDS = (
    "issue_id",
    "original_issue_id",
    "machine_number",
    "logged_by",
    "title",
    "description",
    "severity",
    "category",
    "created_at_utc",
    "resolved_at_utc",
    "resolved_by",
    "solution",
)
REQUIRED_FIELDS = ("machine_number", "title", "created_at_utc", "resolved_at_utc")


class ArchiveSearchError(RuntimeError):
    pass


class DeepArchiveSearchSignals(QObject):
    started = Signal()
    batch_found = Signal(list)
    progress_updated = Signal(str)
    failed = Signal(str)
    cancelled = Signal()
    finished = Signal(int)


class DeepArchiveSearchTask(QRunnable):
    def __init__(
        self,
        archive_path: Path,
        *,
        query: str,
        state_filter: str,
        machine_number: str = "",
        existing_keys: Iterable[str] = (),
        limit: int = DEEP_SEARCH_LIMIT,
    ):
        super().__init__()
        self.archive_path = archive_path
        self.query = query
        self.state_filter = state_filter
        self.machine_number = machine_number
        self.existing_keys = set(existing_keys)
        self.limit = limit
        self.signals = DeepArchiveSearchSignals()
        self._cancelled = False

    def cancel(self) -> None:
        self._cancelled = True

    @Slot()
    def run(self) -> None:
        if self.state_filter == "open":
            self.signals.finished.emit(0)
            return
        self.signals.started.emit()
        emitted = 0

        def emit_batch(batch: list[IssueSearchResult]) -> None:
            nonlocal emitted
            emitted += len(batch)
            self.signals.batch_found.emit(batch)

        try:
            search_excel_archive(
                self.archive_path,
                query=self.query,
                machine_number=self.machine_number,
                existing_keys=self.existing_keys,
                limit=self.limit,
                batch_callback=emit_batch,
                progress_callback=self.signals.progress_updated.emit,
                cancel_requested=lambda: self._cancelled,
            )
        except Exception as exc:
            logger.exception("Deep Search could not read Excel archive at %s", self.archive_path)
            self.signals.failed.emit(str(exc))
            return

        if self._cancelled:
            self.signals.cancelled.emit()
            return
        self.signals.finished.emit(emitted)


def search_excel_archive(
    archive_path: Path,
    *,
    query: str,
    machine_number: str = "",
    existing_keys: Iterable[str] = (),
    limit: int = DEEP_SEARCH_LIMIT,
    batch_callback: Callable[[list[IssueSearchResult]], None] | None = None,
    progress_callback: Callable[[str], None] | None = None,
    cancel_requested: Callable[[], bool] | None = None,
) -> list[IssueSearchResult]:
    if not archive_path.exists():
        raise ArchiveSearchError("Archive workbook is missing.")

    started_at = perf_now()
    perf_log("excel_access", operation="deep_search.load_workbook", path=archive_path)
    terms = _query_terms(query)
    max_results = max(0, min(int(limit), DEEP_SEARCH_HARD_LIMIT))
    seen = set(existing_keys)
    results: list[IssueSearchResult] = []
    batch: list[IssueSearchResult] = []
    workbook = None
    try:
        workbook = load_workbook(archive_path, read_only=True, data_only=True)
        perf_log("excel_access", operation="deep_search.load_workbook_done", path=archive_path, elapsed_ms=elapsed_ms(started_at))
        if ARCHIVE_SHEET not in workbook.sheetnames:
            raise ArchiveSearchError(f"Archive workbook is missing sheet {ARCHIVE_SHEET!r}.")
        worksheet = workbook[ARCHIVE_SHEET]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration as exc:
            raise ArchiveSearchError("Archive workbook has no header row.") from exc
        header_map = _header_map(headers)
        missing = [field for field in REQUIRED_FIELDS if field not in header_map]
        if missing:
            raise ArchiveSearchError(f"Archive workbook is missing required column(s): {', '.join(missing)}.")

        for scanned, row in enumerate(rows, start=1):
            if cancel_requested is not None and cancel_requested():
                break
            if progress_callback is not None and scanned % 250 == 0:
                progress_callback(f"Searching full archive... {scanned} rows scanned")
            record = _record_from_row(row, header_map)
            if machine_number and _text(record.get("machine_number")) != machine_number:
                continue
            if terms and not _record_matches(record, terms):
                continue
            result = _result_from_record(record)
            dedupe_key = search_result_dedupe_key(result)
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            results.append(result)
            batch.append(result)
            if batch_callback is not None and len(batch) >= EXCEL_BATCH_SIZE:
                batch_callback(batch)
                batch = []
            if len(results) >= max_results:
                break
    finally:
        if workbook is not None:
            workbook.close()

    if batch_callback is not None and batch:
        batch_callback(batch)
    perf_log("deep_search.excel_complete", results=len(results), elapsed_ms=elapsed_ms(started_at))
    return results


def search_result_dedupe_key(result: IssueSearchResult) -> str:
    public_issue_id = _text(result.public_issue_id)
    if public_issue_id:
        return f"issue-id:{public_issue_id.casefold()}"
    return "|".join(
        (
            "fallback",
            _text(result.machine_number).casefold(),
            _text(result.title).casefold(),
            _text(result.created_at),
            _text(result.resolved_at),
        )
    )


def _header_map(headers: Iterable[Any]) -> dict[str, int]:
    mapped: dict[str, int] = {}
    for index, header in enumerate(headers):
        key = _text(header).strip().casefold()
        if key:
            mapped[key] = index
    return mapped


def _record_from_row(row: tuple[Any, ...], header_map: dict[str, int]) -> dict[str, Any]:
    record: dict[str, Any] = {}
    for field, index in header_map.items():
        record[field] = row[index] if index < len(row) else None
    return record


def _record_matches(record: dict[str, Any], terms: list[str]) -> bool:
    search_text = " ".join(_text(record.get(field)).casefold() for field in SEARCH_FIELDS)
    return all(term in search_text for term in terms)


def _result_from_record(record: dict[str, Any]) -> IssueSearchResult:
    public_issue_id = _text(record.get("issue_id")) or _text(record.get("original_issue_id"))
    return IssueSearchResult(
        state="excel",
        source="Excel Archive",
        issue_id=_int_or_zero(record.get("cache_id")) or _int_or_zero(record.get("original_issue_id")),
        public_issue_id=public_issue_id,
        machine_number=_text(record.get("machine_number")),
        machine_name=_text(record.get("machine_number")),
        machine_model="",
        title=_text(record.get("title")),
        description=_text(record.get("description")),
        status=_text(record.get("severity")),
        category=_text(record.get("category")),
        logged_by=_text(record.get("logged_by")),
        created_at=_text(record.get("created_at_utc")),
        updated_at=_text(record.get("resolved_at_utc")),
        resolved_at=_text(record.get("resolved_at_utc")),
        resolved_by=_text(record.get("resolved_by")),
        resolution=_text(record.get("solution")),
        history_text="",
    )


def _query_terms(query: str) -> list[str]:
    return [term for term in " ".join(query.casefold().split()).split(" ") if term]


def _text(value: Any) -> str:
    return str(value or "").strip()


def _int_or_zero(value: Any) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0
