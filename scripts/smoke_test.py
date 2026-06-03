from __future__ import annotations

import tempfile
from pathlib import Path
import sys

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
APP_DIR = PROJECT_ROOT / "app"
for path in (APP_DIR, PROJECT_ROOT):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from beeline_issue_tracker.data.archive import ARCHIVE_SHEET, GROUPED_SHEET, ExcelArchive
from beeline_issue_tracker.data.database import initialize_database
from beeline_issue_tracker.data.repository import IssueRepository
from beeline_issue_tracker.domain import LINE_DOWN, NON_CRITICAL, NO_ISSUES


DEMO_MACHINES = (
    ("DEMO-101", "Demo Molder 101", "Demo Hive", "Cell A", "DEMO-ASSET-101", 10),
    ("DEMO-102", "Demo Press 102", "Demo Hive", "Cell B", "DEMO-ASSET-102", 20),
)


def main() -> int:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        db_path = root / "beeline.sqlite3"
        archive_path = root / ".archive" / "archive.xlsx"

        initialize_database(db_path, DEMO_MACHINES)
        repository = IssueRepository(db_path)

        machines = repository.list_machines_with_status()
        assert len(machines) >= 1
        machine_number = machines[0].machine_number
        assert machines[0].calculated_status == NO_ISSUES

        repository.log_issue(
            machine_number=machine_number,
            logged_by="Smoke Test",
            title="Minor sensor drift",
            description="Sensor value is drifting but line is still running.",
            severity=NON_CRITICAL,
            category="Sensor",
        )
        repository.log_issue(
            machine_number=machine_number,
            logged_by="Smoke Test",
            title="Guard open",
            description="Machine stopped because guard switch is open.",
            severity=LINE_DOWN,
            category="Safety",
        )
        summary = repository.get_machine_summary(machine_number)
        assert summary is not None
        assert summary.calculated_status == LINE_DOWN
        assert summary.open_issue_count == 2

        active = repository.list_active_issues(machine_number)
        resolved = repository.resolve_issue(
            active[0].id,
            solution="Closed guard and verified switch.",
            resolved_by="Smoke Test",
        )
        ExcelArchive(archive_path).append_resolved_issue(resolved)
        repository.mark_archive_result(resolved.id, success=True)

        summary = repository.get_machine_summary(machine_number)
        assert summary is not None
        assert summary.calculated_status == NON_CRITICAL
        assert summary.open_issue_count == 1
        assert repository.list_recent_resolved_issues(machine_number)[0].archive_status == "archived"
        assert archive_path.exists()
        workbook = load_workbook(archive_path, data_only=True)
        assert ARCHIVE_SHEET in workbook.sheetnames
        assert GROUPED_SHEET in workbook.sheetnames
        worksheet = workbook[ARCHIVE_SHEET]
        assert worksheet.sheet_state == "visible"
        assert worksheet.max_row == 2
        assert worksheet["A2"].value == resolved.id
        assert worksheet["E2"].value == resolved.title
        grouped = workbook[GROUPED_SHEET]
        assert grouped["A1"].value == "BeeLine Resolved Issues by Date"
        assert grouped.row_dimensions[5].outlineLevel == 1

    print("BeeLine smoke test passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
