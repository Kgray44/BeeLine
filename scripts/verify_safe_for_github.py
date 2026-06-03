from __future__ import annotations

import json
import os
import re
import sqlite3
import subprocess
from pathlib import Path
import sys

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
APP_DIR = PROJECT_ROOT / "app"
for path in (APP_DIR, PROJECT_ROOT):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from beeline_issue_tracker.data.archive import ARCHIVE_SHEET, GROUPED_SHEET, HEADERS


RUNTIME_DIRS = {"data", "archive", "backups", "logs", "config"}
SAFE_RUNTIME_PLACEHOLDERS = {
    "data/.gitkeep",
    "data/README.md",
    "archive/.gitkeep",
    "archive/README.md",
    "backups/.gitkeep",
    "backups/README.md",
    "logs/.gitkeep",
    "logs/README.md",
    "config/.gitkeep",
    "config/README.md",
}
ALLOWED_TEMPLATES = {
    "templates/beeline_config.template.json",
    "templates/beeline.template.sqlite",
    "templates/beeline_archive.template.xlsx",
}
ALLOWED_BINARY_TEMPLATES = {
    "templates/beeline.template.sqlite",
    "templates/beeline_archive.template.xlsx",
}
ALLOWED_MEDIA_PLACEHOLDERS = {
    "assets/branding/nolato_logo_placeholder.png",
}
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".tiff"}
SECRET_NAME_RE = re.compile(
    r"(^|[/\\])(\.env|.*secret.*|.*token.*|.*password.*|.*credential.*|.*apikey.*|.*api_key.*)$",
    re.IGNORECASE,
)
SENSITIVE_NAME_RE = re.compile(
    r"(badge|employee|operator|user|personnel|payroll|plant[_-]?data|machine[_-]?export)",
    re.IGNORECASE,
)
EMAIL_RE = re.compile(r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b", re.IGNORECASE)
PRIVATE_LINK_RE = re.compile(
    r"https?://[^\s)\"']*(sharepoint\.com|1drv\.ms|onedrive\.live\.com|teams\.microsoft\.com)",
    re.IGNORECASE,
)
GITHUB_TOKEN_RE = re.compile(r"\bgh[pousr]_[A-Za-z0-9_]{20,}\b")
OPENAI_KEY_RE = re.compile(r"\bsk-(?:proj-)?[A-Za-z0-9_-]{20,}\b")
AWS_KEY_RE = re.compile(r"\b(?:AKIA|ASIA)[A-Z0-9]{16}\b")
PRIVATE_KEY_RE = re.compile(r"-----BEGIN (?:RSA |DSA |EC |OPENSSH )?PRIVATE KEY-----")
CREDENTIAL_ASSIGNMENT_RE = re.compile(
    r"\b(?:api[_-]?key|access[_-]?token|auth[_-]?token|github[_-]?token|openai[_-]?api[_-]?key|"
    r"password|passwd|pwd|secret|credential)\b\s*[:=]\s*[\"']?[A-Za-z0-9_./+=:@!#$%^&*()-]{8,}",
    re.IGNORECASE,
)
WINDOWS_USER_PATH_RE = re.compile(r"\b[A-Z]:\\Users\\[^\\\s]+\\", re.IGNORECASE)
BADGE_VALUE_RE = re.compile(
    r"\b(?:badge|employee\s*id|operator\s*id)\b[^\n]{0,30}\b\d{3,}\b",
    re.IGNORECASE,
)
PLANT_EXPORT_RE = re.compile(
    r"\b(?:real[-_\s]+)?(?:plant|machine)[-_\s]*(?:export|extract|dump)s?\b",
    re.IGNORECASE,
)
PREDICTIVE_REPORT_RE = re.compile(
    r"(^|[/\\])(?:predictive|machine_predictive).*(?:summary|report).*\.txt$",
    re.IGNORECASE,
)
REAL_DATA_RE = re.compile(
    r"\breal\s+(?:machine|plant|employee|operator)\s+data\b",
    re.IGNORECASE,
)


def main() -> int:
    failures: list[str] = []
    warnings: list[str] = []

    existing_files = _existing_files()
    staged_or_tracked = _git_files("--cached") | _git_files()

    _check_existing_sensitive_files(existing_files, failures)
    _check_git_sensitive_files(staged_or_tracked, failures)
    _check_git_file_contents(staged_or_tracked, failures)
    _check_templates(failures)

    if warnings:
        print("Warnings:")
        for warning in warnings:
            print(f"  - {warning}")

    if failures:
        print("BeeLine GitHub safety check FAILED:")
        for failure in failures:
            print(f"  - {failure}")
        return 1

    print("BeeLine GitHub safety check passed.")
    print("Runtime data, templates, config, archive, logs, and obvious secrets look safe for commit.")
    return 0


def _existing_files() -> set[str]:
    files: set[str] = set()
    skip_dirs = {".git", "__pycache__", ".pytest_cache", ".venv", "venv"}
    for root, dirs, filenames in os.walk(PROJECT_ROOT):
        dirs[:] = [dirname for dirname in dirs if dirname not in skip_dirs]
        root_path = Path(root)
        for filename in filenames:
            path = root_path / filename
            rel = _rel(path)
            if rel.endswith(".pyc"):
                continue
            files.add(rel)
    return files


def _git_files(*args: str) -> set[str]:
    command = ["git", "ls-files", *args]
    try:
        result = subprocess.run(
            command,
            cwd=PROJECT_ROOT,
            check=False,
            capture_output=True,
            text=True,
        )
    except FileNotFoundError:
        return set()
    if result.returncode != 0:
        return set()
    return {line.strip().replace("\\", "/") for line in result.stdout.splitlines() if line.strip()}


def _check_existing_sensitive_files(files: set[str], failures: list[str]) -> None:
    for rel in sorted(files):
        if rel in ALLOWED_TEMPLATES or rel in SAFE_RUNTIME_PLACEHOLDERS or rel in ALLOWED_MEDIA_PLACEHOLDERS:
            continue
        suffix = Path(rel).suffix.lower()
        if suffix in {".sqlite", ".sqlite3", ".db", ".xlsx", ".xlsm", ".xls"}:
            if _is_allowed_runtime_path(rel):
                continue
            failures.append(f"Sensitive data-like file exists outside ignored runtime/template folders: {rel}")
        if suffix in {".log", ".zip", ".bak"} and not _is_allowed_runtime_path(rel):
            failures.append(f"Log or backup file exists outside ignored runtime folders: {rel}")
        if PREDICTIVE_REPORT_RE.search(rel) and not _is_allowed_runtime_path(rel):
            failures.append(f"Generated predictive report exists outside ignored runtime folders: {rel}")
        if SECRET_NAME_RE.search(rel) and rel not in ALLOWED_TEMPLATES:
            failures.append(f"Secret-like file is present: {rel}")


def _check_git_sensitive_files(files: set[str], failures: list[str]) -> None:
    for rel in sorted(files):
        if rel in ALLOWED_TEMPLATES or rel in SAFE_RUNTIME_PLACEHOLDERS or rel in ALLOWED_MEDIA_PLACEHOLDERS:
            continue
        parts = rel.split("/")
        suffix = Path(rel).suffix.lower()
        if parts[0] in {"screenshots", "captures", "exports", "attachments"}:
            failures.append(f"NDA-sensitive generated folder is tracked or staged: {rel}")
        if parts[0] in RUNTIME_DIRS:
            failures.append(f"Runtime file is tracked or staged and must not be committed: {rel}")
        if suffix in {".sqlite", ".sqlite3", ".db", ".xlsx", ".xlsm", ".xls", ".log", ".zip", ".bak"}:
            failures.append(f"Sensitive file type is tracked or staged: {rel}")
        if PREDICTIVE_REPORT_RE.search(rel):
            failures.append(f"Generated predictive report is tracked or staged: {rel}")
        if suffix in IMAGE_SUFFIXES:
            failures.append(f"Image/media file is tracked or staged outside the placeholder whitelist: {rel}")
        if SECRET_NAME_RE.search(rel):
            failures.append(f"Secret-like file is tracked or staged: {rel}")
        if SENSITIVE_NAME_RE.search(rel) and not rel.startswith("app/") and not rel.startswith("tests/"):
            failures.append(f"Sensitive-name file is tracked or staged: {rel}")


def _check_git_file_contents(files: set[str], failures: list[str]) -> None:
    for rel in sorted(files):
        if rel in ALLOWED_BINARY_TEMPLATES or rel in ALLOWED_MEDIA_PLACEHOLDERS:
            continue
        path = PROJECT_ROOT / rel
        if not path.exists() or not path.is_file() or _is_binary_file(path):
            continue
        try:
            text = path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            continue
        for issue in scan_text_for_sensitive_content(rel, text):
            failures.append(issue)


def scan_text_for_sensitive_content(rel: str, text: str) -> list[str]:
    rel = rel.replace("\\", "/")
    findings: list[str] = []
    checks = (
        (EMAIL_RE, "Email address"),
        (PRIVATE_LINK_RE, "SharePoint/OneDrive/Teams link"),
        (GITHUB_TOKEN_RE, "GitHub token"),
        (OPENAI_KEY_RE, "OpenAI-style API key"),
        (AWS_KEY_RE, "AWS access key"),
        (PRIVATE_KEY_RE, "Private key block header"),
        (CREDENTIAL_ASSIGNMENT_RE, "Credential or token assignment"),
        (WINDOWS_USER_PATH_RE, "Windows user profile path"),
        (BADGE_VALUE_RE, "Badge/employee/operator ID value"),
        (PLANT_EXPORT_RE, "Generated export/extract reference"),
    )
    for pattern, label in checks:
        if pattern.search(text):
            findings.append(f"{label} found in tracked/staged text file: {rel}")

    if not _is_doc_file(rel) and REAL_DATA_RE.search(text) and not _is_safe_warning_text(text):
        findings.append(f"Real plant/machine data wording found in non-doc file: {rel}")
    return findings


def _check_templates(failures: list[str]) -> None:
    config_template = PROJECT_ROOT / "templates" / "beeline_config.template.json"
    sqlite_template = PROJECT_ROOT / "templates" / "beeline.template.sqlite"
    archive_template = PROJECT_ROOT / "templates" / "beeline_archive.template.xlsx"

    for path in (config_template, sqlite_template, archive_template):
        if not path.exists():
            failures.append(f"Missing required safe template: {_rel(path)}")

    if config_template.exists():
        try:
            config_text = config_template.read_text(encoding="utf-8")
            config = json.loads(config_text)
            machines = config.get("machines", [])
            if machines:
                failures.append("Config template should not contain real machine rows; keep machines empty or fake.")
            for issue in scan_text_for_sensitive_content(_rel(config_template), config_text):
                if "Real plant/machine data wording" not in issue:
                    failures.append(issue)
        except Exception as exc:
            failures.append(f"Config template is not valid JSON: {exc}")

    if sqlite_template.exists():
        try:
            conn = sqlite3.connect(sqlite_template)
            try:
                expected_tables = (
                    "machines",
                    "active_issues",
                    "resolved_issues_cache",
                    "issue_events",
                    "issue_attachments",
                    "predictive_alerts",
                )
                existing_tables = {
                    row[0]
                    for row in conn.execute(
                        "SELECT name FROM sqlite_master WHERE type = 'table'"
                    ).fetchall()
                }
                table_counts = {
                    table: conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
                    for table in expected_tables
                    if table in existing_tables
                }
            finally:
                conn.close()
            for table, count in table_counts.items():
                if count:
                    failures.append(f"SQLite template table {table} contains {count} records; expected zero.")
        except Exception as exc:
            failures.append(f"Could not validate SQLite template: {exc}")

    if archive_template.exists():
        try:
            workbook = load_workbook(archive_template, read_only=True, data_only=True)
            try:
                if ARCHIVE_SHEET not in workbook.sheetnames:
                    failures.append(f"Excel template is missing {ARCHIVE_SHEET!r}.")
                if GROUPED_SHEET not in workbook.sheetnames:
                    failures.append(f"Excel template is missing {GROUPED_SHEET!r}.")

                if ARCHIVE_SHEET in workbook.sheetnames:
                    worksheet = workbook[ARCHIVE_SHEET]
                    header = tuple(cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1)))
                    if header[: len(HEADERS)] != HEADERS:
                        failures.append("Excel template archive headers do not match BeeLine archive schema.")
                    data_rows = [
                        row for row in worksheet.iter_rows(min_row=2, values_only=True)
                        if any(value is not None for value in row)
                    ]
                    if data_rows:
                        failures.append(f"Excel template contains {len(data_rows)} archive records; expected zero.")
            finally:
                workbook.close()
        except Exception as exc:
            failures.append(f"Could not validate Excel archive template: {exc}")


def _is_allowed_runtime_path(rel: str) -> bool:
    first = rel.split("/", 1)[0]
    return first in RUNTIME_DIRS


def _is_binary_file(path: Path) -> bool:
    try:
        chunk = path.read_bytes()[:4096]
    except OSError:
        return True
    return b"\0" in chunk


def _is_doc_file(rel: str) -> bool:
    rel = rel.replace("\\", "/").casefold()
    return rel.endswith(".md") or rel.startswith("docs/") or rel.endswith("/readme.md")


def _is_safe_warning_text(text: str) -> bool:
    lowered = " ".join(text.casefold().split())
    warning_phrases = (
        "never commit real machine data",
        "do not commit real machine data",
        "real plant runtime data belongs in ignored local folders",
        "config template should not contain real machine rows",
    )
    return any(phrase in lowered for phrase in warning_phrases)


def _rel(path: Path) -> str:
    return path.resolve().relative_to(PROJECT_ROOT.resolve()).as_posix()


if __name__ == "__main__":
    raise SystemExit(main())
