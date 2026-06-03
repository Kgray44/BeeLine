from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from beeline_issue_tracker.domain import LINE_DOWN, NON_CRITICAL, NO_ISSUES, UNKNOWN_ERROR, ResolvedIssue, display_issue_id


ARCHIVE_SHEET = "Resolved_Issues"
GROUPED_SHEET = "Resolved_By_Date"
INFO_SHEET = "Archive_Info"
HEADERS = (
    "cache_id",
    "original_issue_id",
    "issue_id",
    "machine_number",
    "logged_by",
    "title",
    "description",
    "severity",
    "category",
    "created_at_utc",
    "resolved_at_utc",
    "resolved_by",
    "solution",
)
LEGACY_HEADERS = tuple(header for header in HEADERS if header != "issue_id")
GROUPED_HEADERS = (
    "Resolved Time",
    "Issue ID",
    "Machine",
    "Status When Logged",
    "Issue Title",
    "Problem Description",
    "Solution",
    "Logged By",
    "Logged At",
    "Resolved By",
    "Resolved At",
    "Time Open",
    "Category",
    "Notes",
)
GROUPED_REFRESH_ROW_THRESHOLD = 500


DARK_FILL = "1D252C"
DARK_SUBTLE_FILL = "2B3540"
AMBER_FILL = "F6B73C"
AMBER_LIGHT_FILL = "FFF1CC"
ROW_FILL = "FFFFFF"
ROW_ALT_FILL = "F7FAFC"
BORDER_COLOR = "D9E0E6"
TEXT_DARK = "1D252C"
TEXT_LIGHT = "FFFFFF"
MUTED_TEXT = "59636B"
LINE_DOWN_FILL = "F8D7DA"
LINE_DOWN_TEXT = "9F1D20"
NON_CRITICAL_FILL = "FFF1B8"
NON_CRITICAL_TEXT = "6B4B00"
NO_ISSUES_FILL = "D9F2E4"
NO_ISSUES_TEXT = "17613A"
UNKNOWN_FILL = "E5E7EB"
UNKNOWN_TEXT = "4B5563"


@dataclass(frozen=True)
class ArchiveWriteResult:
    archive_path: Path
    sheet_name: str
    row_number: int
    created_workbook: bool
    grouped_refresh_deferred: bool = False


@dataclass(frozen=True)
class ArchiveInspection:
    archive_path: Path
    exists: bool
    sheet_name: str
    row_count: int
    latest_cache_id: int | None
    latest_resolved_at: str
    latest_title: str
    sheet_names: tuple[str, ...] = ()
    grouped_sheet_exists: bool = False
    error: str = ""


class ExcelArchive:
    def __init__(self, archive_path: Path):
        self.archive_path = archive_path

    def append_resolved_issue(self, issue: ResolvedIssue) -> ArchiveWriteResult:
        self.archive_path.parent.mkdir(parents=True, exist_ok=True)
        created_workbook = not self.archive_path.exists()
        workbook = self._load_or_create_workbook()
        raw_sheet = self._get_archive_sheet(workbook)
        row_number = self._append_raw_issue_if_missing(raw_sheet, issue)
        record_count = _raw_record_count(raw_sheet)
        defer_grouped = record_count > GROUPED_REFRESH_ROW_THRESHOLD
        self._refresh_workbook(workbook, defer_grouped=defer_grouped)
        workbook.save(self.archive_path)
        return ArchiveWriteResult(
            archive_path=self.archive_path,
            sheet_name=ARCHIVE_SHEET,
            row_number=row_number,
            created_workbook=created_workbook,
            grouped_refresh_deferred=defer_grouped,
        )

    def _load_or_create_workbook(self):
        if self.archive_path.exists():
            workbook = load_workbook(self.archive_path)
        else:
            workbook = Workbook()
            workbook.active.title = INFO_SHEET
        self._ensure_workbook_layout(workbook)
        return workbook

    def _ensure_workbook_layout(self, workbook) -> None:
        self._get_info_sheet(workbook)
        self._get_archive_sheet(workbook)

    def _get_info_sheet(self, workbook):
        if INFO_SHEET in workbook.sheetnames:
            info = workbook[INFO_SHEET]
        else:
            info = workbook.create_sheet(INFO_SHEET, 0)
        info.sheet_state = "visible"
        return info

    def _get_archive_sheet(self, workbook):
        if ARCHIVE_SHEET not in workbook.sheetnames:
            worksheet = workbook.create_sheet(ARCHIVE_SHEET)
            self._initialize_archive_sheet(worksheet)
        else:
            worksheet = workbook[ARCHIVE_SHEET]
            self._ensure_headers(worksheet)
        worksheet.sheet_state = "visible"
        self._style_raw_sheet(worksheet)
        return worksheet

    def _append_raw_issue_if_missing(self, worksheet, issue: ResolvedIssue) -> int:
        existing_rows: dict[int, int] = {}
        for row_index, (existing_id,) in enumerate(
            worksheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True),
            start=2,
        ):
            if existing_id is None:
                continue
            existing_rows[int(existing_id)] = row_index
        if issue.id in existing_rows:
            return existing_rows[issue.id]

        worksheet.append(
            (
                issue.id,
                issue.original_issue_id,
                display_issue_id(issue),
                issue.machine_number,
                issue.logged_by,
                issue.title,
                issue.description,
                issue.severity,
                issue.category,
                issue.created_at,
                issue.resolved_at,
                issue.resolved_by,
                issue.solution,
            )
        )
        self._style_raw_sheet(worksheet)
        return worksheet.max_row

    def _refresh_workbook(self, workbook, *, defer_grouped: bool = False) -> None:
        raw_sheet = self._get_archive_sheet(workbook)
        records = _raw_records(raw_sheet)
        self._refresh_info_sheet(
            self._get_info_sheet(workbook),
            len(records),
            grouped_refresh_deferred=defer_grouped,
        )
        if not defer_grouped:
            self._refresh_grouped_sheet(workbook, records)

    def _refresh_info_sheet(
        self,
        worksheet,
        total_resolved: int,
        *,
        grouped_refresh_deferred: bool = False,
    ) -> None:
        _clear_sheet(worksheet)
        worksheet.title = INFO_SHEET
        worksheet.sheet_view.showGridLines = False

        worksheet.merge_cells("A1:D1")
        worksheet["A1"] = "BeeLine Resolved Issue Archive"
        worksheet["A1"].font = Font(bold=True, size=20, color=TEXT_LIGHT)
        worksheet["A1"].fill = PatternFill(fill_type="solid", fgColor=DARK_FILL)
        worksheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[1].height = 32

        worksheet.merge_cells("A3:D3")
        worksheet["A3"] = (
            "Resolved issue records are stored in this workbook. "
            "The BeeLine kiosk UI does not read this workbook during normal use."
        )
        worksheet["A3"].alignment = Alignment(wrap_text=True, vertical="top")
        worksheet["A3"].font = Font(color=TEXT_DARK)
        worksheet.row_dimensions[3].height = 42

        rows = (
            ("Archive updated UTC", _utc_now_iso()),
            ("Total resolved rows", total_resolved),
            ("Raw source sheet", ARCHIVE_SHEET),
            ("Readable grouped view", GROUPED_SHEET),
            (
                "Grouped refresh",
                "Deferred; run python run_beeline.py --repair-archive"
                if grouped_refresh_deferred
                else "Current",
            ),
            ("Archive writer", "openpyxl, no Excel COM required"),
        )
        start_row = 5
        for offset, (label, value) in enumerate(rows):
            row = start_row + offset
            worksheet.cell(row=row, column=1, value=label)
            worksheet.cell(row=row, column=2, value=value)
            worksheet.cell(row=row, column=1).font = Font(bold=True, color=TEXT_DARK)
            worksheet.cell(row=row, column=2).font = Font(color=TEXT_DARK)
            for column in range(1, 3):
                cell = worksheet.cell(row=row, column=column)
                cell.border = _thin_border()
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.fill = PatternFill(fill_type="solid", fgColor=ROW_ALT_FILL if offset % 2 else ROW_FILL)

        worksheet.column_dimensions["A"].width = 26
        worksheet.column_dimensions["B"].width = 44
        worksheet.column_dimensions["C"].width = 18
        worksheet.column_dimensions["D"].width = 18

    def _refresh_grouped_sheet(self, workbook, records: list[dict[str, Any]]) -> None:
        if GROUPED_SHEET in workbook.sheetnames:
            del workbook[GROUPED_SHEET]

        raw_index = workbook.sheetnames.index(ARCHIVE_SHEET) if ARCHIVE_SHEET in workbook.sheetnames else 0
        worksheet = workbook.create_sheet(GROUPED_SHEET, raw_index + 1)
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A4"
        worksheet.sheet_properties.outlinePr.summaryBelow = False
        worksheet.sheet_properties.outlinePr.summaryRight = False

        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(GROUPED_HEADERS))
        title = worksheet.cell(row=1, column=1, value="BeeLine Resolved Issues by Date")
        title.font = Font(bold=True, size=18, color=TEXT_LIGHT)
        title.fill = PatternFill(fill_type="solid", fgColor=DARK_FILL)
        title.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[1].height = 30

        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(GROUPED_HEADERS))
        subtitle = worksheet.cell(
            row=2,
            column=1,
            value=f"Newest resolved dates are shown first. Updated {_utc_now_iso()}.",
        )
        subtitle.font = Font(color=MUTED_TEXT)
        subtitle.fill = PatternFill(fill_type="solid", fgColor=ROW_FILL)

        for column, header in enumerate(GROUPED_HEADERS, start=1):
            cell = worksheet.cell(row=3, column=column, value=header)
            cell.font = Font(bold=True, color=TEXT_LIGHT)
            cell.fill = PatternFill(fill_type="solid", fgColor=DARK_SUBTLE_FILL)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _thin_border()

        row_index = 4
        grouped = _group_records_by_resolved_date(records)
        for date_key in sorted(grouped.keys(), reverse=True):
            day_records = sorted(grouped[date_key], key=_resolved_sort_key, reverse=True)
            row_index = self._write_date_group(worksheet, row_index, date_key, day_records)

        if row_index == 4:
            worksheet.cell(row=row_index, column=1, value="No resolved issues archived yet.")
            worksheet.cell(row=row_index, column=1).font = Font(italic=True, color=MUTED_TEXT)

        worksheet.auto_filter.ref = f"A3:N{max(row_index - 1, 3)}"
        _set_grouped_widths(worksheet)

    def _write_date_group(self, worksheet, row_index: int, date_key, records: list[dict[str, Any]]) -> int:
        line_down_count = sum(1 for record in records if record.get("severity") == LINE_DOWN)
        non_critical_count = sum(1 for record in records if record.get("severity") == NON_CRITICAL)

        header_row = row_index
        header_cells = (
            _date_label(date_key),
            f"{len(records)} issue{'s' if len(records) != 1 else ''}",
            f"{line_down_count} line down",
            f"{non_critical_count} non-critical",
        )
        for column, value in enumerate(header_cells, start=1):
            cell = worksheet.cell(row=header_row, column=column, value=value)
            cell.font = Font(bold=True, color=TEXT_DARK)
            cell.fill = PatternFill(fill_type="solid", fgColor=AMBER_FILL)
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center")
        for column in range(5, len(GROUPED_HEADERS) + 1):
            cell = worksheet.cell(row=header_row, column=column)
            cell.fill = PatternFill(fill_type="solid", fgColor=AMBER_FILL)
            cell.border = _thin_border()
        worksheet.row_dimensions[header_row].height = 24

        first_issue_row = header_row + 1
        row_index = first_issue_row
        for offset, record in enumerate(records):
            self._write_grouped_issue_row(worksheet, row_index, record, offset)
            row_index += 1

        last_issue_row = row_index - 1
        if last_issue_row >= first_issue_row:
            worksheet.row_dimensions.group(first_issue_row, last_issue_row, outline_level=1, hidden=False)
        return row_index

    def _write_grouped_issue_row(self, worksheet, row_index: int, record: dict[str, Any], offset: int) -> None:
        resolved_dt = _parse_dt(record.get("resolved_at_utc"))
        created_dt = _parse_dt(record.get("created_at_utc"))
        values = (
            _time_label(resolved_dt, record.get("resolved_at_utc")),
            record.get("issue_id") or record.get("original_issue_id") or record.get("cache_id"),
            record.get("machine_number"),
            record.get("severity"),
            record.get("title"),
            record.get("description"),
            record.get("solution"),
            record.get("logged_by"),
            record.get("created_at_utc"),
            record.get("resolved_by"),
            record.get("resolved_at_utc"),
            _time_open(created_dt, resolved_dt),
            record.get("category"),
            "",
        )
        fill = ROW_ALT_FILL if offset % 2 else ROW_FILL
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_index, column=column, value=value)
            cell.fill = PatternFill(fill_type="solid", fgColor=fill)
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="top", wrap_text=column in {5, 6, 7, 14})
            cell.font = Font(color=TEXT_DARK)

        status_cell = worksheet.cell(row=row_index, column=4)
        status_fill, status_text = _status_style(str(status_cell.value or UNKNOWN_ERROR))
        status_cell.fill = PatternFill(fill_type="solid", fgColor=status_fill)
        status_cell.font = Font(bold=True, color=status_text)

    @staticmethod
    def _initialize_archive_sheet(worksheet) -> None:
        worksheet.append(HEADERS)
        ExcelArchive._style_raw_sheet(worksheet)

    @staticmethod
    def _ensure_headers(worksheet) -> None:
        first_row_values = [cell.value for cell in worksheet[1][: len(HEADERS)]]
        if all(value is None for value in first_row_values):
            for index, header in enumerate(HEADERS, start=1):
                worksheet.cell(row=1, column=index, value=header)
            ExcelArchive._style_raw_sheet(worksheet)
            return
        if tuple(first_row_values) != HEADERS:
            legacy_values = tuple(cell.value for cell in worksheet[1][: len(LEGACY_HEADERS)])
            if legacy_values == LEGACY_HEADERS:
                worksheet.insert_cols(3)
                worksheet.cell(row=1, column=3, value="issue_id")
                for row in range(2, worksheet.max_row + 1):
                    original_issue_id = worksheet.cell(row=row, column=2).value
                    worksheet.cell(row=row, column=3, value=str(original_issue_id or ""))
                ExcelArchive._style_raw_sheet(worksheet)
                return
            raise ValueError(
                f"Archive sheet {ARCHIVE_SHEET!r} has unexpected headers. "
                "Refusing to append to a workbook with a different layout."
            )
        ExcelArchive._style_raw_sheet(worksheet)

    @staticmethod
    def _style_raw_sheet(worksheet) -> None:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:M{max(worksheet.max_row, 1)}"
        widths = {
            "A": 12,
            "B": 18,
            "C": 18,
            "D": 16,
            "E": 18,
            "F": 32,
            "G": 48,
            "H": 16,
            "I": 18,
            "J": 24,
            "K": 24,
            "L": 18,
            "M": 48,
        }
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width

        for cell in worksheet[1][: len(HEADERS)]:
            cell.font = Font(bold=True, color=TEXT_LIGHT)
            cell.fill = PatternFill(fill_type="solid", fgColor=DARK_FILL)
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in range(2, worksheet.max_row + 1):
            row_fill = ROW_ALT_FILL if (row - 2) % 2 else ROW_FILL
            for column in range(1, len(HEADERS) + 1):
                cell = worksheet.cell(row=row, column=column)
                cell.fill = PatternFill(fill_type="solid", fgColor=row_fill)
                cell.border = _thin_border()
                cell.alignment = Alignment(vertical="top", wrap_text=column in {6, 7, 13})
                cell.font = Font(color=TEXT_DARK)
            status_cell = worksheet.cell(row=row, column=8)
            status_fill, status_text = _status_style(str(status_cell.value or UNKNOWN_ERROR))
            status_cell.fill = PatternFill(fill_type="solid", fgColor=status_fill)
            status_cell.font = Font(bold=True, color=status_text)


def create_empty_archive_workbook(archive_path: Path) -> None:
    archive_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.active.title = INFO_SHEET
    archive = ExcelArchive(archive_path)
    archive._ensure_workbook_layout(workbook)
    archive._refresh_workbook(workbook)
    workbook.save(archive_path)


def refresh_archive_workbook(archive_path: Path) -> None:
    archive_path.parent.mkdir(parents=True, exist_ok=True)
    archive = ExcelArchive(archive_path)
    workbook = archive._load_or_create_workbook()
    archive._refresh_workbook(workbook)
    workbook.save(archive_path)


def inspect_archive(archive_path: Path) -> ArchiveInspection:
    if not archive_path.exists():
        return ArchiveInspection(
            archive_path=archive_path,
            exists=False,
            sheet_name=ARCHIVE_SHEET,
            row_count=0,
            latest_cache_id=None,
            latest_resolved_at="",
            latest_title="",
            sheet_names=(),
            grouped_sheet_exists=False,
        )

    workbook = None
    try:
        workbook = load_workbook(archive_path, read_only=True, data_only=True)
        sheet_names = tuple(workbook.sheetnames)
        if ARCHIVE_SHEET not in workbook.sheetnames:
            return ArchiveInspection(
                archive_path=archive_path,
                exists=True,
                sheet_name=ARCHIVE_SHEET,
                row_count=0,
                latest_cache_id=None,
                latest_resolved_at="",
                latest_title="",
                sheet_names=sheet_names,
                grouped_sheet_exists=GROUPED_SHEET in workbook.sheetnames,
                error=f"Workbook is missing expected sheet {ARCHIVE_SHEET!r}.",
            )

        worksheet = workbook[ARCHIVE_SHEET]
        records = _raw_records(worksheet)
        latest = max(records, key=_resolved_sort_key) if records else None
        return ArchiveInspection(
            archive_path=archive_path,
            exists=True,
            sheet_name=ARCHIVE_SHEET,
            row_count=len(records),
            latest_cache_id=int(latest["cache_id"]) if latest and latest.get("cache_id") is not None else None,
            latest_resolved_at=str(latest["resolved_at_utc"]) if latest else "",
            latest_title=str(latest["title"]) if latest else "",
            sheet_names=sheet_names,
            grouped_sheet_exists=GROUPED_SHEET in workbook.sheetnames,
        )
    except Exception as exc:
        return ArchiveInspection(
            archive_path=archive_path,
            exists=True,
            sheet_name=ARCHIVE_SHEET,
            row_count=0,
            latest_cache_id=None,
            latest_resolved_at="",
            latest_title="",
            sheet_names=(),
            grouped_sheet_exists=False,
            error=str(exc),
        )
    finally:
        if workbook is not None:
            workbook.close()


def _raw_records(worksheet) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    headers = _raw_headers(worksheet)
    for row in worksheet.iter_rows(min_row=2, max_col=len(headers), values_only=True):
        if not any(value is not None for value in row):
            continue
        record = dict(zip(headers, row))
        if "issue_id" not in record:
            record["issue_id"] = str(record.get("original_issue_id") or "")
        records.append(record)
    return records


def _raw_record_count(worksheet) -> int:
    count = 0
    for row in worksheet.iter_rows(min_row=2, max_col=len(HEADERS), values_only=True):
        if any(value is not None for value in row):
            count += 1
    return count


def _raw_headers(worksheet) -> tuple[str, ...]:
    first_row_values = tuple(cell.value for cell in worksheet[1][: len(HEADERS)])
    if first_row_values == HEADERS:
        return HEADERS
    legacy_values = tuple(cell.value for cell in worksheet[1][: len(LEGACY_HEADERS)])
    if legacy_values == LEGACY_HEADERS:
        return LEGACY_HEADERS
    return HEADERS


def _group_records_by_resolved_date(records: list[dict[str, Any]]):
    grouped = defaultdict(list)
    for record in records:
        resolved_dt = _parse_dt(record.get("resolved_at_utc"))
        key = resolved_dt.date() if resolved_dt else _fallback_date(record.get("resolved_at_utc"))
        grouped[key].append(record)
    return grouped


def _resolved_sort_key(record: dict[str, Any]) -> float:
    resolved_dt = _parse_dt(record.get("resolved_at_utc"))
    if resolved_dt is None:
        return 0.0
    if resolved_dt.tzinfo is None:
        resolved_dt = resolved_dt.replace(tzinfo=timezone.utc)
    return resolved_dt.timestamp()


def _parse_dt(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if not value:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def _fallback_date(value: Any):
    parsed = _parse_dt(value)
    if parsed:
        return parsed.date()
    text = str(value or "").strip()
    if len(text) >= 10:
        try:
            return datetime.fromisoformat(text[:10]).date()
        except ValueError:
            pass
    return datetime.min.date()


def _date_label(date_key) -> str:
    if hasattr(date_key, "isoformat"):
        return date_key.isoformat()
    return str(date_key)


def _time_label(resolved_dt: datetime | None, fallback: Any) -> str:
    if resolved_dt is None:
        return str(fallback or "")
    return resolved_dt.strftime("%H:%M:%S")


def _time_open(created_dt: datetime | None, resolved_dt: datetime | None) -> str:
    if created_dt is None or resolved_dt is None:
        return ""
    if created_dt.tzinfo is None:
        created_dt = created_dt.replace(tzinfo=timezone.utc)
    if resolved_dt.tzinfo is None:
        resolved_dt = resolved_dt.replace(tzinfo=timezone.utc)
    seconds = max(0, int((resolved_dt - created_dt).total_seconds()))
    days, rem = divmod(seconds, 86400)
    hours, rem = divmod(rem, 3600)
    minutes, _ = divmod(rem, 60)
    if days:
        return f"{days}d {hours}h {minutes}m"
    if hours:
        return f"{hours}h {minutes}m"
    return f"{minutes}m"


def _status_style(status: str) -> tuple[str, str]:
    if status == LINE_DOWN:
        return LINE_DOWN_FILL, LINE_DOWN_TEXT
    if status == NON_CRITICAL:
        return NON_CRITICAL_FILL, NON_CRITICAL_TEXT
    if status == NO_ISSUES:
        return NO_ISSUES_FILL, NO_ISSUES_TEXT
    return UNKNOWN_FILL, UNKNOWN_TEXT


def _set_grouped_widths(worksheet) -> None:
    widths = {
        "A": 15,
        "B": 18,
        "C": 14,
        "D": 18,
        "E": 28,
        "F": 42,
        "G": 42,
        "H": 18,
        "I": 24,
        "J": 18,
        "K": 24,
        "L": 15,
        "M": 18,
        "N": 24,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def _clear_sheet(worksheet) -> None:
    if worksheet.max_row:
        worksheet.delete_rows(1, worksheet.max_row)
    if worksheet.max_column:
        worksheet.delete_cols(1, worksheet.max_column)


def _thin_border() -> Border:
    side = Side(style="thin", color=BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()
